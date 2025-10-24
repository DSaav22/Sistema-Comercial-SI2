from django.http import HttpResponse
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from datetime import datetime


def generate_pdf_report(data, parsed_data) -> HttpResponse:
    """
    Genera un reporte en formato PDF usando ReportLab.
    
    Args:
        data: Lista de diccionarios con los datos del reporte
        parsed_data: Diccionario con información parseada del prompt
        
    Returns:
        HttpResponse con el PDF
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Título del reporte
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    title = Paragraph("Reporte de Ventas - SmartSales365", title_style)
    elements.append(title)
    
    # Información de fechas
    if parsed_data.get('start_date') or parsed_data.get('end_date'):
        date_info = f"Período: "
        if parsed_data.get('start_date'):
            date_info += f"{parsed_data['start_date'].strftime('%d/%m/%Y')}"
        if parsed_data.get('end_date'):
            if parsed_data.get('start_date'):
                date_info += f" - {parsed_data['end_date'].strftime('%d/%m/%Y')}"
            else:
                date_info += f"Hasta {parsed_data['end_date'].strftime('%d/%m/%Y')}"
        
        date_paragraph = Paragraph(date_info, styles['Normal'])
        elements.append(date_paragraph)
        elements.append(Spacer(1, 12))
    
    # Fecha de generación
    gen_date = Paragraph(
        f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        styles['Normal']
    )
    elements.append(gen_date)
    elements.append(Spacer(1, 20))
    
    if not data:
        no_data = Paragraph("No se encontraron datos para el período especificado.", styles['Normal'])
        elements.append(no_data)
    else:
        # Crear tabla con los datos
        table_data = []
        
        # Encabezados
        if data:
            headers = list(data[0].keys())
            table_data.append(headers)
            
            # Datos
            for row in data:
                table_data.append([str(row.get(key, '')) for key in headers])
        
        # Crear la tabla
        table = Table(table_data)
        
        # Estilo de la tabla
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ])
        
        table.setStyle(table_style)
        elements.append(table)
        
        # Resumen al final
        elements.append(Spacer(1, 20))
        summary = Paragraph(
            f"Total de registros: {len(data)}",
            styles['Normal']
        )
        elements.append(summary)
    
    # Construir el PDF
    doc.build(elements)
    
    # Preparar la respuesta
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_ventas.pdf"'
    
    return response


def generate_excel_report(data, parsed_data) -> HttpResponse:
    """
    Genera un reporte en formato Excel usando openpyxl.
    
    Args:
        data: Lista de diccionarios con los datos del reporte
        parsed_data: Diccionario con información parseada del prompt
        
    Returns:
        HttpResponse con el archivo Excel
    """
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Reporte de Ventas"
    
    # Estilos
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    title_font = Font(name='Arial', size=16, bold=True)
    title_alignment = Alignment(horizontal='center', vertical='center')
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    worksheet.merge_cells('A1:E1')
    worksheet['A1'] = 'Reporte de Ventas - SmartSales365'
    worksheet['A1'].font = title_font
    worksheet['A1'].alignment = title_alignment
    
    # Información de fechas
    row_num = 2
    if parsed_data.get('start_date') or parsed_data.get('end_date'):
        date_info = "Período: "
        if parsed_data.get('start_date'):
            date_info += f"{parsed_data['start_date'].strftime('%d/%m/%Y')}"
        if parsed_data.get('end_date'):
            if parsed_data.get('start_date'):
                date_info += f" - {parsed_data['end_date'].strftime('%d/%m/%Y')}"
            else:
                date_info += f"Hasta {parsed_data['end_date'].strftime('%d/%m/%Y')}"
        
        worksheet.merge_cells(f'A{row_num}:E{row_num}')
        worksheet[f'A{row_num}'] = date_info
        row_num += 1
    
    # Fecha de generación
    worksheet.merge_cells(f'A{row_num}:E{row_num}')
    worksheet[f'A{row_num}'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    row_num += 2
    
    if not data:
        worksheet[f'A{row_num}'] = "No se encontraron datos para el período especificado."
    else:
        # Encabezados
        headers = list(data[0].keys())
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = header.replace('_', ' ').title()
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        row_num += 1
        
        # Datos
        for row_data in data:
            for col_num, header in enumerate(headers, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = row_data.get(header, '')
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = border
            row_num += 1
        
        # Ajustar ancho de columnas
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column].width = adjusted_width
        
        # Resumen
        row_num += 1
        worksheet[f'A{row_num}'] = f"Total de registros: {len(data)}"
        worksheet[f'A{row_num}'].font = Font(bold=True)
    
    # Guardar en memoria
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    
    # Preparar la respuesta
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="reporte_ventas.xlsx"'
    
    return response
